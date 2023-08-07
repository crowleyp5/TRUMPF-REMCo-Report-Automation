import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
import win32com.client as win32

spreadsheet1 = '***************'
spreadsheet2 = '******************'
output_file = '**************************'

df1 = pd.read_csv(spreadsheet1)
df2 = pd.read_csv(spreadsheet2)
df1.to_csv(spreadsheet1, index=False)
df2.to_csv(spreadsheet2, index=False)

# Clean the spreadsheets
df1 = df1.dropna(how='all')
df2 = df2.dropna(how='all')
df1 = df1.drop(df1.columns[-2:], axis=1)
df2 = df2.drop(df2.columns[-2:], axis=1)
df2 = df2[(df2.iloc[:, 11] != '8. Complete')]
df1.iloc[:, 12] = df1.iloc[:, 12].fillna('')
df2.iloc[:, 12] = df2.iloc[:, 12].fillna('')
df1.iloc[:, 15] = df1.iloc[:, 15].fillna('')
df2.iloc[:, 15] = df2.iloc[:, 15].fillna('')
df1.iloc[:, 14] = df1.iloc[:, 14].astype(str).replace('[^0-9.]', '', regex=True)
df2.iloc[:, 14] = df2.iloc[:, 14].astype(str).replace('[^0-9.]', '', regex=True)
df1.iloc[:, 12] = df1.iloc[:, 12].astype(str).replace('[^a-zA-Z0-9_]', '', regex=True)
df2.iloc[:, 12] = df2.iloc[:, 12].astype(str).replace('[^a-zA-Z0-9_]', '', regex=True)

marked_cells_row = np.array([])
marked_cells_column = np.array([])
workbook = load_workbook(output_file)
sheet = workbook.active

# Find matching rows between the new and old versions
for i in range(len(df2)):
    serial_number = df2.iloc[i, 6]
    matching_row = df1[df1.iloc[:, 6] == serial_number]
    if not matching_row.empty:
        for j in range(len(df2.iloc[i, :])):
            if df2.iloc[i, j] != matching_row.iloc[0, j]:
                marked_cells_row = np.append(marked_cells_row, i + 2)
                marked_cells_column = np.append(marked_cells_column, j + 1)
            df2.iloc[i, j] = matching_row.iloc[0, j]

# Mark row indices of lasers present in the new version but not in the old one as blue
blue_rows = np.array([])
for k in range(len(df1)):
    serial_number = df1.iloc[k, 6]
    matching_row = df2[df2.iloc[:, 6] == serial_number]
    not_complete = False
    if df1.iloc[k, 11] in ['1. New order', '2. Order Issues', '3. In-Process', '4. Confirmed', '5. Approved', '6. Shipped', '7. Delivered']:
        not_complete = True
    if matching_row.empty and not_complete:
        df2.loc[-1] = df1.iloc[k, :]
        blue_rows = np.append(blue_rows, (k + 2))

# Mark the row indices with subsequent deliveries as beige
beige_rows = np.array([])
for l in range(len(df2)):
    if '9' in df2.iloc[l, 11] or '10' in df2.iloc[l, 11]:
        beige_rows = np.append(beige_rows, (l + 2))

# Mark the rows indices that have just been delivered as grey
grey_rows = np.array([])
for m in range(len(df2)):
    if '8' in df2.iloc[m, 11]:
        grey_rows = np.append(grey_rows, (m + 2))

# Clear the spreadsheets
sheet.delete_rows(1, sheet.max_row)
for row in sheet.iter_rows():
    for cell in row:
        cell.value = None
        cell.fill = None

# Set custom column widths
column_widths = [19, 20, 18, 12, 15, 23, 13, 16, 23, 14, 10, 30, 18, 17, 24, 52]
for i, width in enumerate(column_widths):
    column_letter = get_column_letter(i + 1)
    sheet.column_dimensions[column_letter].width = width

# Set headers
headers = df2.columns.tolist()
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font = Font(color='FFFFFF', name='Arial', bold=True)
sheet.append(headers)
for cell in sheet[1]:
    cell.fill = header_fill
    cell.font = header_font

# Input cell values
for _, row in df2.iterrows():
    row_values = [cell if not pd.isnull(cell) else "" for cell in row.tolist()]
    sheet.append(row_values)

# Highlight all cells within the dimensions of df2 green
green_fill = PatternFill(start_color='A9D08E', end_color='A9D08E', fill_type='solid')
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
    for cell in row:
        cell.fill = green_fill

# Highlight the beige rows
beige_fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
for row_index in beige_rows:
    row_index = int(row_index)
    for cell in sheet[row_index]:
        cell.fill = beige_fill

# Highlight the grey rows
grey_fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
for row_index in grey_rows:
    row_index = int(row_index)
    for cell in sheet[row_index]:
        cell.fill = grey_fill

# Highlight the rows in blue_rows blue
blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
for row_index in blue_rows:
    row_index = int(row_index)
    for cell in sheet.iter_rows(min_row=row_index, max_row=row_index, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.fill = blue_fill

# Highlight the marked cells with yellow fill, excluding the 16th column
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
for row_index, column_index in zip(marked_cells_row, marked_cells_column):
    row_index = int(row_index)
    column_index = int(column_index)
    if column_index != 16:  # Exclude the 16th column
        cell = sheet.cell(row=row_index, column=column_index)
        cell.fill = yellow_fill

# Set the default zoom level to 75%
sheet.sheet_view.zoomScale = 85

# Add black lined borders to all cells
border = Border(left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000'))
for row in range(2, sheet.max_row + 1):
    row_values = [cell.value for cell in sheet[row]]
    if any(row_values):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            cell.border = border

# Add hyperlinks to tracking column
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=13, max_col=13):
    for cell in row:
        if cell.value:
            print(cell.value)
            display_text = cell.value
            print(display_text)
            hawb = display_text[-10:]
            print(hawb)
            link = 'https://www.ups.com/track?loc=en_US&requester=QUIC&tracknum=' + hawb + '/trackdetails'
            cell.value = '=HYPERLINK("{}", "{}")'.format(link, display_text)

workbook.save(output_file)

# Open the Excel file
excel = win32.gencache.EnsureDispatch('Excel.Application')
wb = excel.Workbooks.Open('****************')
ws = wb.ActiveSheet

# Define the range to be sorted (including all columns)
sort_range = ws.Range(ws.Cells(2, 1), ws.Cells(ws.UsedRange.Rows.Count, ws.UsedRange.Columns.Count))

# Sort the range based on the second column in ascending order
sort_range.Sort(Key1=ws.Cells(2, 2), Order1=1, Header=1, Orientation=1)

wb.Save()
wb.Close()
excel.Quit()
